"""Генерация .xlsx-отчётов через openpyxl. RU/EN/TJ."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .i18n import t

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 40) -> None:
    """Подбирает ширину колонок по содержимому. Учёт многобайтовой кириллицы
    через len(str(...)); для xlsx это норм, чисел не считаем."""
    for col_idx, col in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(longest + 2, 10), max_width
        )


def _write_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"  # шапка зафиксирована при скролле
    ws.row_dimensions[1].height = 22


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_iso(s: Any) -> datetime | None:
    """Возвращает naive datetime — Excel не поддерживает tz-aware значения.
    DSS отдаёт `+05:00`, для отчёта важно локальное wall-time, поэтому
    отбрасываем tzinfo без конвертации."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s.replace(tzinfo=None) if s.tzinfo else s
    try:
        dt = datetime.fromisoformat(str(s))
    except ValueError:
        return None
    return dt.replace(tzinfo=None) if dt.tzinfo else dt


def _to_min(t_: time) -> int:
    return t_.hour * 60 + t_.minute


def generate_workers_xlsx(
    rows: Iterable[Mapping[str, Any]],
    lang: str = "ru",
) -> bytes:
    """Лист «Работники»: Имя · Проходов · Последний раз · Дверь."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Workers"  # лист именуется латиницей — лимит Excel 31, без локализации
    headers = [
        t("csv_w_name", lang),
        t("csv_w_passes", lang),
        t("csv_w_last_seen", lang),
        t("csv_w_last_door", lang),
    ]
    _write_headers(ws, headers)
    for r in rows:
        last_seen = _parse_iso(r.get("last_seen"))
        ws.append([
            (r.get("person_name") or "").strip(),
            int(r.get("total_passes") or 0),
            last_seen,  # openpyxl сам отдаст как datetime
            (r.get("last_door") or "").strip(),
        ])
    # Формат даты для столбца C (last_seen)
    for cell in ws["C"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    _autosize(ws)
    return _save(wb)


def generate_teachers_template_xlsx(lang: str = "ru") -> bytes:
    """Шаблон для импорта учителей: 6 колонок + 2 строки-примера.

    Колонки совпадают с парсером `parse_teachers_xlsx`. Примеры окрашены
    серым курсивом, чтобы пользователь сразу видел, что это образец, а не
    реальные данные."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    headers = [
        "Subject",
        "Teacher's Name (en)",
        "Имя учителя (ru)",
        "Номи омӯзгор (tg)",
        "Phone Number",
        "Position",
    ]
    _write_headers(ws, headers)

    examples = [
        [
            "забони англисӣ",
            "Sharon Olowey",
            "Шарон Олоуэй",
            "Шарон Олоуэй",
            "+992900000001",
            "учитель",
        ],
        [
            "Табиатшиносӣ",
            "Kurbonova Umeda",
            "Курбонова Умеда",
            "Қурбонова Умеда",
            "+992900000002",
            "учитель",
        ],
        [
            "",
            "Bobokhonov Karim",
            "Бобохонов Карим",
            "Бобохонов Карим",
            "+992900000003",
            "Methodist",
        ],
    ]
    grey_italic = Font(italic=True, color="808080")
    for row in examples:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = grey_italic

    _autosize(ws)
    return _save(wb)


def generate_absent_xlsx(
    regulars: Iterable[Mapping[str, Any]],
    seen_by_day: Mapping[str, set],
    start_date: date,
    end_date: date,
    lang: str = "ru",
) -> tuple[bytes, int]:
    """Pivot-лист «Absent»: №, ФИО, [день1] [день2] ..., Дней не было.

    В ячейке — «✗» если человек не пришёл в этот рабочий день. Воскресенья
    пропускаются. В строки попадают только те, кто пропустил минимум один
    день. Сортировка — по числу пропусков убыванию.

    Возвращает (xlsx_bytes, n_rows): n_rows — сколько людей попало в отчёт
    (нужно вызывающему, чтобы решить, слать файл или текст «никого нет»).
    """
    days: list[date] = []
    cur = start_date
    while cur <= end_date:
        if cur.weekday() != 6:
            days.append(cur)
        cur += timedelta(days=1)

    seen = {k: set(str(p) for p in v) for k, v in seen_by_day.items()}

    items: list[tuple[str, str, list[date]]] = []  # (pid, name, absent_days)
    for r in regulars:
        pid = str(r.get("person_id") or "").strip()
        if not pid:
            continue
        name = (r.get("person_name") or pid).strip() or pid
        absent_days: list[date] = []
        for d in days:
            if pid not in seen.get(d.isoformat(), set()):
                absent_days.append(d)
        if absent_days:
            items.append((pid, name, absent_days))

    # Сорт: по числу пропусков убыванию, имя по алфавиту
    items.sort(key=lambda x: (-len(x[2]), x[1].lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Absent"

    headers: list[str] = [t("late_h_no", lang), t("csv_h_name", lang)]
    headers.extend(d.strftime("%d.%m") for d in days)
    headers.append(t("absent_h_days", lang))
    _write_headers(ws, headers)

    for idx, (_pid, name, absent_days) in enumerate(items, start=1):
        absent_set = {d.isoformat() for d in absent_days}
        row: list[Any] = [idx, name]
        for d in days:
            row.append("✗" if d.isoformat() in absent_set else "")
        row.append(len(absent_days))
        ws.append(row)

    if items:
        ws.append([])
        footer: list[Any] = [t("late_xlsx_total", lang), ""]
        for d in days:
            cnt = sum(
                1 for _, _, ad in items
                if d.isoformat() in {x.isoformat() for x in ad}
            )
            footer.append(cnt if cnt else "")
        footer.append(len(items))
        ws.append(footer)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    _autosize(ws)
    return _save(wb), len(items)


def generate_trend_xlsx(
    daily: list[Mapping[str, Any]],
    lang: str = "ru",
) -> bytes:
    """Лист «Trend»: Дата · Пришли · Опоздали · Не пришли + LineChart.

    `daily` — отсортированный по дате список словарей с ключами
    {date: 'YYYY-MM-DD', came, late, absent}. Воскресенья на стороне
    вызывающего отфильтровываются.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Trend"

    headers = [
        t("mon_trend_h_date", lang),
        t("mon_trend_h_came", lang),
        t("mon_trend_h_late", lang),
        t("mon_trend_h_absent", lang),
    ]
    _write_headers(ws, headers)

    for r in daily:
        ws.append([
            r.get("date"),
            int(r.get("came") or 0),
            int(r.get("late") or 0),
            int(r.get("absent") or 0),
        ])

    if daily:
        n = len(daily)
        chart = LineChart()
        chart.title = t("mon_trend_chart", lang)
        chart.y_axis.title = ""
        chart.x_axis.title = ""
        chart.height = 10
        chart.width = 22
        data_ref = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=n + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "F2")

    _autosize(ws)
    return _save(wb)


def generate_top_late_xlsx(
    rows: list[Mapping[str, Any]],
    lang: str = "ru",
) -> bytes:
    """Лист «TopLate»: №, ФИО, Отдел, Должность, Дней опоздал, Всего (мин),
    В среднем (мин/день).

    `rows` — уже отсортированный список словарей с ключами
    {name, dept, position, days_late, total_min, avg_min}.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TopLate"

    headers = [
        t("mon_top_h_rank", lang),
        t("mon_top_h_name", lang),
        t("mon_top_h_dept", lang),
        t("mon_top_h_pos", lang),
        t("mon_top_h_days", lang),
        t("mon_top_h_total", lang),
        t("mon_top_h_avg", lang),
    ]
    _write_headers(ws, headers)

    for idx, r in enumerate(rows, start=1):
        ws.append([
            idx,
            r.get("name") or "",
            r.get("dept") or "—",
            r.get("position") or "—",
            int(r.get("days_late") or 0),
            int(r.get("total_min") or 0),
            round(float(r.get("avg_min") or 0), 1),
        ])

    _autosize(ws)
    return _save(wb)


def generate_late_xlsx(
    rows: Iterable[Mapping[str, Any]],
    work_start: time,
    start_date: date,
    end_date: date,
    lang: str = "ru",
) -> bytes:
    """Pivot-лист «Late»: №, ФИО, [день1] [день2] ..., Итог (мин), Дней.

    Колонки-даты — все рабочие дни диапазона [start_date..end_date],
    воскресенья пропускаются. В ячейке — минуты опоздания за этот день
    (пусто, если человек не опоздал или не пришёл). Сортировка по сумме
    опоздания убыванию. Внизу — строка с суммами по каждому дню.
    """
    work_start_min = _to_min(work_start)

    # Список рабочих дней (skip Sunday=6)
    days: list[date] = []
    cur = start_date
    while cur <= end_date:
        if cur.weekday() != 6:
            days.append(cur)
        cur += timedelta(days=1)

    # person_id (или имя) → {day_iso: late_min}
    by_person: dict[str, dict[str, int]] = defaultdict(dict)
    name_by_pid: dict[str, str] = {}

    for r in rows:
        first_in = _parse_iso(r.get("first_in"))
        if first_in is None:
            continue
        in_m = first_in.hour * 60 + first_in.minute
        late = in_m - work_start_min
        if late <= 0:
            continue
        day_raw = (r.get("day") or first_in.date().isoformat() or "")[:10]
        try:
            day_d = date.fromisoformat(day_raw)
        except ValueError:
            continue
        if day_d.weekday() == 6:
            continue
        if day_d < start_date or day_d > end_date:
            continue
        name = (r.get("person_name") or "").strip()
        pid = str(r.get("person_id") or name or "").strip()
        if not pid:
            continue
        if name and pid not in name_by_pid:
            name_by_pid[pid] = name
        elif pid not in name_by_pid:
            name_by_pid[pid] = pid
        # Если в одном дне несколько проходов — берём максимум опоздания.
        prev = by_person[pid].get(day_d.isoformat(), 0)
        by_person[pid][day_d.isoformat()] = max(prev, int(late))

    person_totals = {pid: sum(d.values()) for pid, d in by_person.items()}
    persons_sorted = sorted(
        by_person.keys(),
        key=lambda p: (-person_totals[p], name_by_pid.get(p, "").lower()),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Late"

    headers: list[str] = [t("late_h_no", lang), t("csv_h_name", lang)]
    headers.extend(d.strftime("%d.%m") for d in days)
    headers.extend([t("late_h_summary", lang), t("late_h_days", lang)])
    _write_headers(ws, headers)

    for idx, pid in enumerate(persons_sorted, start=1):
        per_day = by_person[pid]
        row: list[Any] = [idx, name_by_pid[pid]]
        days_late = 0
        total_min = 0
        for d in days:
            v = per_day.get(d.isoformat())
            if v:
                row.append(int(v))
                days_late += 1
                total_min += int(v)
            else:
                row.append("")
        row.append(total_min)
        row.append(days_late)
        ws.append(row)

    if persons_sorted:
        ws.append([])  # разделитель
        footer: list[Any] = [t("late_xlsx_total", lang), ""]
        grand_total = 0
        for d in days:
            col_sum = sum(
                by_person[pid].get(d.isoformat(), 0)
                for pid in persons_sorted
            )
            footer.append(col_sum if col_sum else "")
            grand_total += col_sum
        footer.append(grand_total)
        footer.append(len(persons_sorted))
        ws.append(footer)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    _autosize(ws)
    return _save(wb)


def generate_attendance_xlsx(
    rows: Iterable[Mapping[str, Any]],
    work_start: time,
    work_end: time,
    lang: str = "ru",
) -> bytes:
    """Лист «Отчёт»: Дата · Имя · Первый вход · Последний выход ·
    Опоздание · Ранний уход · Проходов · Статус."""
    work_start_min = _to_min(work_start)
    work_end_min = _to_min(work_end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    headers = [
        t("csv_h_date", lang),
        t("csv_h_name", lang),
        t("csv_h_in", lang),
        t("csv_h_out", lang),
        t("csv_h_late", lang),
        t("csv_h_early", lang),
        t("csv_h_passes", lang),
        t("csv_h_status", lang),
    ]
    _write_headers(ws, headers)

    for r in rows:
        day = r.get("day") or ""
        name = (r.get("person_name") or "").strip()
        first_in = _parse_iso(r.get("first_in"))
        last_out = _parse_iso(r.get("last_out"))
        passes = int(r.get("passes") or 0)

        late_val: int | str = ""
        early_val: int | str = ""
        status = ""

        if first_in is None:
            status = t("csv_st_only_out", lang)
        else:
            in_m = first_in.hour * 60 + first_in.minute
            if in_m >= work_end_min:
                status = t("csv_st_after_hours", lang)
            elif in_m > work_start_min:
                late_val = in_m - work_start_min
                status = t("csv_st_late", lang)
            else:
                if last_out is None:
                    status = t("csv_st_no_out", lang)
                else:
                    out_m = last_out.hour * 60 + last_out.minute
                    if out_m < work_end_min:
                        early_val = work_end_min - out_m
                        status = t("csv_st_early", lang)
                    else:
                        status = t("csv_st_on_time", lang)

        ws.append([
            day,
            name,
            first_in,  # время ячейки — datetime
            last_out,
            late_val,
            early_val,
            passes,
            status,
        ])

    # Форматы времени для колонок C и D
    for col in ("C", "D"):
        for cell in ws[col][1:]:
            if cell.value is not None:
                cell.number_format = "hh:mm:ss"

    _autosize(ws)
    return _save(wb)
