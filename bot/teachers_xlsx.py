"""Парсер .xlsx со списком учителей.

Ожидаемая структура:
    A: Subject (предмет)
    B: Teacher's Name (en)
    C: Имя учителя (ru)
    D: Номи омӯзгор (tg)
    E: Phone Number
    F: Position (должность) — опционально, для совместимости со старыми
       файлами без 6-й колонки.

Первая строка — заголовки, пропускается. Пустые строки тоже пропускаются.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import load_workbook


def _norm(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm_phone(v) -> str | None:
    """Сохраняет '+' и цифры, остальное удаляет."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    keep = [c for c in s if c.isdigit() or c == "+"]
    out = "".join(keep)
    return out or None


def parse_teachers_xlsx(data: bytes) -> list[dict]:
    """Возвращает список учителей. Бросает ValueError при пустом/битом файле."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Файл не содержит листов.")
    rows = ws.iter_rows(min_row=2, values_only=True)
    out: list[dict] = []
    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Защита от файла с разным числом колонок: добиваем None до 6.
        cells = list(row)
        if len(cells) < 6:
            cells = cells + [None] * (6 - len(cells))
        subject = _norm(cells[0])
        name_en = _norm(cells[1])
        name_ru = _norm(cells[2])
        name_tg = _norm(cells[3])
        phone = _norm_phone(cells[4])
        position = _norm(cells[5])
        # Минимум одно из имён должно быть, иначе строка бесполезна.
        if not (name_en or name_ru or name_tg):
            continue
        out.append({
            "subject": subject,
            "name_en": name_en,
            "name_ru": name_ru,
            "name_tg": name_tg,
            "phone": phone,
            "position": position,
        })
    if not out:
        raise ValueError("В файле не найдено ни одной строки с учителем.")
    return out
